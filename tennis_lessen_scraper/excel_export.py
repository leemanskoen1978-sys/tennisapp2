"""Export naar Excel met beide bladen: lessen_data en lessen_privaat."""

import logging
import re
from pathlib import Path

from openpyxl import load_workbook, Workbook

from .config import EXCEL_OUTPUT, SHEET_NAME, COLUMNS, IS_CI
from .calendar_config import PRIVATE_LESSONS_COLUMNS, SESSIONS_CALENDAR_ID
from .dates import get_previous_month_range

logger = logging.getLogger(__name__)


def _clean_bedrag(value: str) -> str | int | float:
    """Maak bedrag zuiver cijfer: verwijder €, spaties, normaliseer decimalen."""
    if not value or not str(value).strip():
        return ""
    s = str(value).strip().replace("€", "").replace(" ", "").replace("\xa0", "")
    s = re.sub(r"\s+", "", s)
    if not s:
        return ""
    s = s.replace(",", ".")
    parts = s.split(".")
    if len(parts) == 1:
        try:
            return float(parts[0]) if parts[0] else ""
        except ValueError:
            return value
    dec_part = parts[-1]
    int_part = "".join(parts[:-1])
    try:
        return float(int_part + "." + dec_part)
    except ValueError:
        return value


def append_lessen_to_excel(lessen: list[dict[str, str]]) -> int:
    """Voeg lessen toe aan het tabblad lessen_data (Tennis Vlaanderen website)."""
    if not lessen:
        return 0

    path = Path(EXCEL_OUTPUT)
    existing_data = set()

    if not path.exists():
        if IS_CI:
            wb = Workbook()
            wb.remove(wb.active)
            ws = wb.create_sheet(SHEET_NAME)
            for col_idx, header in enumerate(COLUMNS, 1):
                ws.cell(row=1, column=col_idx, value=header)
            start_row = 2
        else:
            raise FileNotFoundError(f"Excel-bestand niet gevonden: {path}")
    else:
        try:
            wb = load_workbook(path, keep_vba=True)
        except Exception as e:
            logger.warning("keep_vba mislukt, probeer zonder: %s", e)
            wb = load_workbook(path)

        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            for col_idx, header in enumerate(COLUMNS, 1):
                ws.cell(row=1, column=col_idx, value=header)
            start_row = 2
        else:
            ws = wb[SHEET_NAME]
            start_row = ws.max_row + 1
            if start_row == 2 and ws.cell(1, 1).value is None:
                for col_idx, header in enumerate(COLUMNS, 1):
                    ws.cell(row=1, column=col_idx, value=header)
                start_row = 2
            for row in range(2, start_row):
                key_parts = []
                for i, col in enumerate(COLUMNS[:5], 1):
                    val = ws.cell(row=row, column=i).value
                    key_parts.append(str(val) if val else "")
                existing_data.add("|".join(key_parts))

    added = 0
    for lesson in lessen:
        key_parts = [
            lesson.get("Club", ""),
            lesson.get("Aanbod", ""),
            lesson.get("Doelgroep", ""),
            lesson.get("Groep", ""),
            lesson.get("Dag + uur", ""),
        ]
        key = "|".join(key_parts)
        if key in existing_data:
            continue
        existing_data.add(key)

        for col_idx, col_name in enumerate(COLUMNS, 1):
            val = lesson.get(col_name, "")
            if col_name == "Bedrag":
                val = _clean_bedrag(val)
            ws.cell(row=start_row, column=col_idx, value=val)
        start_row += 1
        added += 1

    try:
        wb.save(path)
    except Exception as e:
        logger.error("Opslaan van Excel faalde: %s. Controleer of het bestand niet geopend is.", e)
        raise

    logger.info("Toegevoegd aan Excel hoofdbestand: %d lessen", added)
    return added


def append_private_lessen_to_excel(privaat_lessen: list[dict[str, str]]) -> int:
    """Voeg privé lessen (Google Calendar) toe aan nieuw Excel tabblad."""
    if not privaat_lessen:
        return 0

    path = Path(EXCEL_OUTPUT)
    
    # Gebruik dezelfde logica als lessen_data
    if not path.exists():
        if IS_CI:
            wb = Workbook()
            wb.remove(wb.active)
            ws = wb.create_sheet("lessen_privaat")
            for col_idx, header in enumerate(PRIVATE_LESSONS_COLUMNS, 1):
                ws.cell(row=1, column=col_idx, value=header)
            start_row = 2
        else:
            raise FileNotFoundError(f"Excel-bestand niet gevonden: {path}")
    else:
        try:
            wb = load_workbook(path, keep_vba=True)
        except Exception as e:
            logger.warning("keep_vba mislukt, probeer zonder: %s", e)
            wb = load_workbook(path)

        if "lessen_privaat" not in wb.sheetnames:
            ws = wb.create_sheet("lessen_privaat")
            for col_idx, header in enumerate(PRIVATE_LESSONS_COLUMNS, 1):
                ws.cell(row=1, column=col_idx, value=header)
            start_row = 2
        else:
            ws = wb["lessen_privaat"]
            start_row = ws.max_row + 1
    
    added = 0
    for lesson in privaat_lessen:
        for col_idx, col_name in enumerate(PRIVATE_LESSONS_COLUMNS, 1):
            val = lesson.get(col_name, "")
            ws.cell(row=start_row, column=col_idx, value=val)
        start_row += 1
        added += 1

    try:
        wb.save(path)
    except Exception as e:
        logger.error("Opslaan van Excel faalde: %s. Controleer of het bestand niet geopend is.", e)
        raise

    logger.info("Toegevoegd aan Excel privé blad: %d lessen", added)
    return added